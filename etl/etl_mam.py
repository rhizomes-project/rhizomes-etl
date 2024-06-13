#!/usr/bin/env python

import re
import requests
import sys

from etl.etl_process import BaseETLProcess
from etl.setup import ETLEnv
from etl.tools import RhizomeField
from etl.date_parsers import get_date_first_four

import openpyxl

#
# See Schema Review: https://docs.google.com/document/d/1iZLT6jTaqVDHl49jTIyBfUW1uAQUoX97-kKtKYaFdOI/edit#heading=h.rgi19g8xnr5
#


field_map = {

    "ACCESSNO":                             RhizomeField.ID,
    "TITLE":                                RhizomeField.TITLE,
    "CREATOR":                              RhizomeField.AUTHOR_ARTIST,
    "OBJECT URL":                           RhizomeField.URL,
    "DESCRIP":                              RhizomeField.DESCRIPTION,
    "STERMS":                               RhizomeField.SUBJECTS_TOPIC_KEYWORDS,
    "DATE":                                 RhizomeField.DATE,
    "COLLECTION":                           RhizomeField.SOURCE,
    "OBJECT IMAGE":                         RhizomeField.IMAGES,

    # "Place Holder" values used to temporarily to fill in "real" values:
    "CREATOR2":                             RhizomeField.AUTHOR_ARTIST,
    "MEDIUM":                               RhizomeField.SUBJECTS_TOPIC_KEYWORDS,
    "OBJNAME":                              RhizomeField.SUBJECTS_TOPIC_KEYWORDS,
    "IMAGEFILE":                            RhizomeField.SUBJECTS_TOPIC_KEYWORDS,

}


field_map_keys = list(field_map.keys())


# Map column name to column index in the input
# data (in case the column order changes)
column_indices = {

    # ACCESSNO
    field_map_keys[0]: 0,

    # TITLE
    field_map_keys[1]: 16,

    # CREATOR
    field_map_keys[2]: 3,

    # OBJECT URL
    field_map_keys[3]: 22,

    # DESCRIP
    field_map_keys[4]: 6,

    # STERMS
    field_map_keys[5]: 14,

    # DATE
    field_map_keys[6]: 5,

    # COLLECTION
    field_map_keys[7]: 1,

    # OBJECT IMAGE
    field_map_keys[8]: 23,

    # CREATOR2
    field_map_keys[9]: 4,

    # MEDIUM
    field_map_keys[10]: 9,

    # OBJNAME
    field_map_keys[11]: 11,

    # IMAGEFILE
    field_map_keys[12]: 7,

}



required_values = [
    "ACCESSNO",
    "TITLE",
    "DESCRIP",
    "OBJECT IMAGE"
]

bad_strings = [
    "\n"
]


def is_record_valid(record):

    for required_value in required_values:

        if not record.get(required_value):

            message = f"Missing {required_value}"

            return False, message

    return True, None

def clean_value(value):

    # Is this value on a blank line?
    if not value:

        return None

    # This function can only clean strings.
    if type(value) is not str:

        return value

    for bad_string in bad_strings:

        value = value.replace(bad_string, " ")

    return value.strip()

def parse_date(value):

    if not value:

        return None

    elif type(value) is not str:

        return value

    # remove "ca. " from the date - e.g., "ca. 1960"
    invalid_prefixes = [
        "ca.",
        "c.a.",
        "ca",
    ]

    value = value.lower()

    for invalid_prefix in invalid_prefixes:

        if value.startswith(invalid_prefix):

            value = value[ len(invalid_prefix) : ]
            break

    return clean_value(value=value)

def ParseExcelValue(value):
    # Remove things like "CONCAT()" and "SUBSTITUTE()" from excel cell values.

    if not value:

        return value


    start = value.find('"')

    end = value.find('"', start + 1)

    value = value[ start + 1 : end ]

    return clean_value(value=value)


class SpecialValueHandler():

    def __init__(self):

        self.reset()

    def reset(self):

        self.creator = ""
        self.subject = ""
        self.image_url = ""

    def add_creator(self, value):

        if not value:

            return

        if self.creator:

            self.creator += ". "

        self.creator += value

    def add_subject(self, value):

        if not value:

            return None

        value = clean_value(value=value)

        if self.subject:

            self.subject += ". "

        self.subject += value

        return self.subject

    def add_image_url(self, value):

        if not value:

            return None

        elif "\\" in value:

            value = value.replace("\\", "/")

        # Image url seems to not resolve correctly unless it
        # is all lower-case
        self.image_url += value.lower()

        return self.image_url


special_value_handler = SpecialValueHandler()


def parse_values(field_name, value):

    if field_name.startswith("CREATOR"):

        special_value_handler.add_creator(value=value)

        return { "CREATOR" : special_value_handler.creator }

    elif field_name == "DATE":

        return { "DATE" : parse_date(value=value) }

    elif field_name in [ "STERMS", "MEDIUM", "OBJNAME" ]:

        return { "STERMS": special_value_handler.add_subject(value=value) }

    elif field_name == "OBJECT URL":

        return { field_name : ParseExcelValue(value=value) }

    elif field_name == "OBJECT IMAGE":

        value = ParseExcelValue(value=value)
        return { field_name : special_value_handler.add_image_url(value=value) }

    elif field_name == "IMAGEFILE":

        return { field_name : special_value_handler.add_image_url(value=value) }

    else:

        return { field_name: clean_value(value=value) }


class MAMETLProcess(BaseETLProcess):

    def init_testing(self):

        pass

    def get_field_map(self):

        return field_map

    def get_collection_name(self):

        return "Mexic-Arte Museum (MAM)"

    def get_access_rights_stmt(self):

        return """Mexic-Arte Museum has created and maintains websites and other digital properties to support its mission to enrich the community through education programs, exhibitions, and interpretations of the collection. These Websites include https://mexic-artemuseum.org/ and https://mexicartemuseum.pastperfectonline.com/. This does not mean that Mexic-Arte Museum owns each component of the compilation, some of which may be owned by others and used with their permission or used in accordance with applicable law (e.g., fair use). Mexic-Arte Museum is committed to protecting the intellectual property rights of visual and performing artists and others who hold copyright. Most items in the collection are protected by copyright and/or related rights. Private study, educational, and non-commercial use of digital images from our websites is permitted, with attribution to the Mexic-Arte Museum. Commercial use of any materials on the Mexic-Arte Museum website is expressly forbidden. Users who wish to obtain permission for publication, display, distribution, or other uses of these materials should contact the rights holder(s)."""

    def get_date_parsers(self):

        return {
            r'^\d{4}':  get_date_first_four
        }

    def extract(self):

        # Open the excel workbook.
        wrkbk = openpyxl.load_workbook("etl/data/permanent/mam/input_20240601.xlsx")
        sheet = wrkbk.active

        data = []

        # Iterate through the sheet by row & column and get the data.
        # Note: row and col numbers in openpyxl are 1-based, and we skip header the row.
        for row_num in range(1, sheet.max_row + 1):

            record = {}
            special_value_handler.reset()

            # Build each row.
            for field_name, col_index in column_indices.items():

                cell_obj = sheet.cell(row=row_num + 1, column=col_index + 1)

                values = parse_values(field_name=field_name, value=cell_obj.value)
                if values:

                    record |= values

            # Finalize the record.
            record["OBJECT IMAGE"] = special_value_handler.image_url

            # REVIEW: Update this once correct value for object url is available.
            record["OBJECT URL"] = None

            # Skip any blank lines.
            is_valid, message = is_record_valid(record=record)
            if is_valid:

                data.append(record)

            else:

                print(f"Row {row_num + 1} is invalid: {message}", file=sys.stderr)

        return data


if __name__ == "__main__":    # pragma: no cover

    from etl.run import run_cmd_line

    args_ = [ "mam" ]

    if len(sys.argv) > 1:

        args_ += sys.argv[1:]

    run_cmd_line(args=args_)
