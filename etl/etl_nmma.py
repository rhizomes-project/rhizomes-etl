#!/usr/bin/env python

import re
import requests
import sys

from etl.etl_process import BaseETLProcess
from etl.setup import ETLEnv
from etl.tools import RhizomeField
from etl.date_parsers import get_date_first_four

import openpyxl


field_map = {

    "accession_num":                           RhizomeField.ID,
    "title":                                   RhizomeField.TITLE,
    "translation":                             RhizomeField.ALTERNATE_TITLES,
    "themes":                                  RhizomeField.SUBJECTS_TOPIC_KEYWORDS,
    "web_url":                                 RhizomeField.URL,
    "creator_accessionPart_full_name":         RhizomeField.AUTHOR_ARTIST,
    "media":                                   RhizomeField.RESOURCE_TYPE,
    "a17dimensions":                           RhizomeField.DIMENSIONS,
    "creation_year":                           RhizomeField.DATE,
    "webcredit":                               RhizomeField.SOURCE,
    "image_url":                               RhizomeField.IMAGES,
    "description":                             RhizomeField.DESCRIPTION,
    "add_to_description_1":                    RhizomeField.DESCRIPTION_ADD_1,
    "add_to_description_2":                    RhizomeField.DESCRIPTION_ADD_2,

}


field_map_keys = list(field_map.keys())


# Map column name to column index in the input
# data (in case NMMA changes column order or
# adds columns)
column_indices = {

    # Accession y Part
    field_map_keys[0]: 0,

    # Title
    field_map_keys[1]: 1,

    # Translation
    field_map_keys[2]: 2,

    # Themes
    field_map_keys[3]: 3,

    # WebAddress
    field_map_keys[4]: 4,

    # Creator~Full Name
    field_map_keys[5]: 5,

    # Media
    field_map_keys[6]: 6,

    # Dimensions
    field_map_keys[7]: 7,

    # Creation Year
    field_map_keys[8]: 8,

    # WebCredit
    field_map_keys[9]: 9,
}



required_values = [
    "accession_num",
    "title",
    "web_url"
]

bad_strings = [
    "\n"
]


def is_record_valid(record):

    for required_value in required_values:

        if not record.get(required_value):

            message = f"Missing {required_value}"

            return False, message

    return True, None

def clean_value(value):

    # Is this value on a blank line?
    if not value:

        return None

    for bad_string in bad_strings:

        value = value.replace(bad_string, " ")

    return value.strip()

def remove_parens(value):
    # Remove parenthesis from beginning and end of value.

    if not value:

        return None

    value = value.strip()

    if value.startswith("("):

        value = value[ 1 : ]

    if value.endswith(")"):

        value = value[ : -1 ]

    return clean_value(value=value)

def parse_title(value):
    # Separate title value into english title
    # and translated title.

    if not value:

        return None

    match =  re.search(r'\(.*\)', value)

    if match:

        title = value[ : match.start() ]
        translation = value[ match.start() : match.end() ]

    else:

        title = value
        translation = None

    results = {
        "title": clean_value(value=title),
        "add_to_description_1": remove_parens(value=translation)
    }

    return results

def parse_alt_title(value):

    if not value:

        return None

    value = remove_parens(value=value)

    # Move text within brackets in translation to
    # "add to description 2"
    match = re.search(r'\[.*\]', value)

    if match:

        translation = value[ : match.start() ]
        add_to_description_2 = value[ match.start() + 1 : match.end() -1 ]

    else:

        translation = value
        add_to_description_2 = None

    return {
        "translation": remove_parens(value=translation),
        "add_to_description_2": add_to_description_2
    }

class DescriptionBuilder():

    def __init__(self):

        self.reset()

    def reset(self):

        self.media = ""
        self.dimensions = ""

    def get_description(self):
        # Description is media + dimension.

        self.media = clean_value(self.media)
        self.dimensions = clean_value(self.dimensions)

        # Capitalize first letter of media.
        self.media = self.media[0].upper() + self.media[1:]

        # Handle periods.
        if not self.media.endswith("."):

            self.media += "."

        self.media += " "

        self.dimensions += "."

        description = self.media + self.dimensions

        self.reset()

        return description


description_builder = DescriptionBuilder()


def parse_subject(value):

    if not value:

        return None

    delimiters = [ ",", ";" ]

    for delimiter in delimiters:

        value = value.replace(delimiter, "|")

    bad_strings = [ " |", "| " ]
    for bad_string in bad_strings:

        value = value.replace(bad_string, "|")

    return {
        "themes": clean_value(value=value)
    }

def parse_values(field_name, value):

    if field_name == "title":

        return parse_title(value=value)

    elif field_name == "translation":

        return parse_alt_title(value=value)

    elif field_name == "themes":

        return parse_subject(value=value)

    elif field_name == "media":

        description_builder.media = value

    elif field_name == "a17dimensions":

        description_builder.dimensions = value

    else:

        return { field_name: clean_value(value=value) }

def get_image_url(record):

    accession_num = record["accession_num"]
    url = record["web_url"]

    response = requests.get(url=url, timeout=60)
    content = response.text

    url_beginning = 'data-srcset="'
    url_end = ".png"

    begin = content.find(url_beginning)
    end = content.find(url_end, begin)

    url = content[ begin + len(url_beginning) : end + len(url_end) ]

    if url:

        return "https://nationalmuseumofmexicanart.org" + url

    else:

        return None


class NMAAETLProcess(BaseETLProcess):

    def init_testing(self):

        pass

    def get_field_map(self):

        return field_map

    def get_output_cols(self):

        return [

            RhizomeField.ID,
            RhizomeField.TITLE,
            RhizomeField.ALTERNATE_TITLES,
            RhizomeField.AUTHOR_ARTIST,
            RhizomeField.IMAGES,
            RhizomeField.URL,
            RhizomeField.DESCRIPTION,
            RhizomeField.DESCRIPTION_ADD_1,
            RhizomeField.DESCRIPTION_ADD_2,
            RhizomeField.SUBJECTS_TOPIC_KEYWORDS,
            RhizomeField.SEARCHABLE_DATE,
            RhizomeField.RESOURCE_TYPE,
            RhizomeField.DIGITAL_FORMAT,
            RhizomeField.SOURCE,
            RhizomeField.LANGUAGE,
            RhizomeField.COLLECTION_NAME,
            RhizomeField.ANNOTATES,
            RhizomeField.ACCESS_RIGHTS,

        ]

    def get_collection_name(self):

        return "National Museum of Mexican Art (NMMA)"

    def get_date_parsers(self):

        # Not: we are just letting all date values come through.

        # return {
        #     r'^\d{4}':  get_date_first_four
        # }

        return None

    def get_access_rights_stmt(self):

        return """NMMA is committed to protecting the intellectual property rights of visual and performing artists and others who hold copyright. With the exception of fair use as defined by US copyright law, NMMA expressly prohibits the reproduction, distribution, downloading, transmission, sale, transfer, creation of derivative works, modification, public display, public performance, or publication of any materials on this website. Commercial use of any materials on the NMMA website is expressly forbidden.\n
Images, text, software, documentation, electronic text and image files, audio and video clips, and other materials (the “Contents”) on this site are either © National Museum of Mexican Art or used with permission by NMMA; and are protected by under United States and international copyright laws.\n
Please note that the NMMA does not hold the copyright to any works in its collection or on exhibition. Therefore, you are responsible for obtaining reproductions rights from any third-party rights holders."""

    def extract(self):

        # Open the excel workbook.
        wrkbk = openpyxl.load_workbook("etl/data/permanent/nmma/NMMA_Proj.Rhizome.xlsx")
        sheet = wrkbk.active

        data = []

        # Iterate through the sheet by row & column and get the data.
        # Note: row and col numbers in openpyxl are 1-based, and we skip header the row.
        for row_num in range(1, sheet.max_row + 1):

            record = {}

            # Build each row.
            for field_name, col_index in column_indices.items():

                cell_obj = sheet.cell(row=row_num + 1, column=col_index + 1)

                values = parse_values(field_name=field_name, value=cell_obj.value)
                if values:

                    record |= values

            # Skip any blank lines.
            is_valid, message = is_record_valid(record=record)
            if is_valid:

                # Scrape the image url from the web page.
                record["image_url"] = get_image_url(record=record)

                # Build the description and add it.
                record["description"] = description_builder.get_description()

                data.append(record)

            else:

                print(f"Row {row_num + 1} is invalid: {message}", file=sys.stderr)

        return data


if __name__ == "__main__":    # pragma: no cover

    from etl.run import run_cmd_line

    args_ = [ "nmma" ]

    if len(sys.argv) > 1:

        args_ += sys.argv[1:]

    run_cmd_line(args=args_)
